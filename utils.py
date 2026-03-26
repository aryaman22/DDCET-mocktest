import csv
import json
import random
import io
from models import db, Question, QuestionBank

def get_bank_names(config):
    """Get comma-separated bank names for an exam config."""
    try:
        bank_ids = json.loads(config.bank_ids) if config.bank_ids else []
    except (json.JSONDecodeError, TypeError):
        bank_ids = []
    
    if not bank_ids:
        return 'None'
        
    banks = QuestionBank.query.filter(QuestionBank.id.in_(bank_ids)).all()
    if not banks:
        return 'None'
        
    return ', '.join(b.name for b in banks)


EXCEL_HEADERS = ['qno', 'section', 'topic', 'question', 'option_a', 'option_b',
                 'option_c', 'option_d', 'correct_ans', 'explanation']


def parse_csv_questions(filepath, bank_id):
    """Parse CSV file and return (list[Question], list[str warnings])."""
    questions = []
    warnings = []

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        # Normalize header names to lowercase
        if reader.fieldnames:
            reader.fieldnames = [h.strip().lower() for h in reader.fieldnames]

        for idx, row in enumerate(reader, start=1):
            row = {k.strip().lower(): (str(v).strip() if v is not None else '') for k, v in row.items() if k}
            q, warn = _validate_row(row, idx, bank_id)
            if q:
                questions.append(q)
            if warn:
                warnings.append(warn)

    return questions, warnings


def parse_json_questions(filepath, bank_id):
    """Parse JSON file and return (list[Question], list[str warnings])."""
    questions = []
    warnings = []

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError as e:
            return [], [f"Invalid JSON file: {e}"]

    if not isinstance(data, list):
        return [], ["JSON must be an array of question objects."]

    for idx, row in enumerate(data, start=1):
        row = {k.strip().lower(): (str(v).strip() if v is not None else '') for k, v in row.items() if k}
        q, warn = _validate_row(row, idx, bank_id)
        if q:
            questions.append(q)
        if warn:
            warnings.append(warn)

    return questions, warnings


def parse_excel_questions(filepath, bank_id):
    """Parse Excel (.xlsx) file and return (list[Question], list[str warnings])."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], ["openpyxl is not installed. Please install it with: pip install openpyxl"]

    questions = []
    warnings = []

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return [], ["Excel file has no active sheet."]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], ["Excel file is empty."]

        raw_headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

        for idx, row_data in enumerate(rows[1:], start=2):
            row = {}
            for col_idx, value in enumerate(row_data):
                if col_idx < len(raw_headers) and raw_headers[col_idx]:
                    row[raw_headers[col_idx]] = str(value).strip() if value is not None else ''

            q, warn = _validate_row(row, idx, bank_id)
            if q:
                questions.append(q)
            if warn:
                warnings.append(warn)

        wb.close()
    except Exception as e:
        return [], [f"Error reading Excel file: {e}"]

    return questions, warnings


def _validate_row(row, idx, bank_id):
    """Validate a single row and return (Question or None, warning or None)."""
    question_text = str(row.get('question', '') or '').strip()
    if not question_text:
        return None, f"Row {idx}: Skipped — empty question text."

    option_a = str(row.get('option_a', '') or '').strip()
    option_b = str(row.get('option_b', '') or '').strip()
    option_c = str(row.get('option_c', '') or '').strip()
    option_d = str(row.get('option_d', '') or '').strip()

    if not all([option_a, option_b, option_c, option_d]):
        return None, f"Row {idx}: Skipped — one or more options are empty."

    correct_ans = str(row.get('correct_ans', '') or '').strip().upper()
    if correct_ans not in ['A', 'B', 'C', 'D', 'X']:
        return None, f"Row {idx}: Skipped — invalid correct_ans '{correct_ans}' (must be A/B/C/D/X)."

    # qno
    try:
        qno = int(row.get('qno', idx))
    except (ValueError, TypeError):
        qno = idx

    # section
    try:
        section = int(row.get('section', 1))
        if section not in [1, 2]:
            section = 1
    except (ValueError, TypeError):
        section = 1

    topic = str(row.get('topic', '') or '').strip() or 'General'
    explanation = str(row.get('explanation', '') or '').strip()

    q = Question(
        bank_id=bank_id,
        qno=qno,
        section=section,
        topic=topic,
        question=question_text,
        option_a=option_a,
        option_b=option_b,
        option_c=option_c,
        option_d=option_d,
        correct_ans=correct_ans,
        explanation=explanation
    )
    return q, None


def select_questions_for_config(config):
    """Select and shuffle questions based on config filters."""
    bank_ids = json.loads(config.bank_ids) if config.bank_ids else []
    if not bank_ids:
        return []

    pool = Question.query.join(QuestionBank).filter(
        QuestionBank.id.in_(bank_ids),
        QuestionBank.is_active == True
    )

    if config.section_filter and config.section_filter != 0:
        pool = pool.filter(Question.section == config.section_filter)

    if config.topic_filter and config.topic_filter.strip():
        topics = [t.strip() for t in config.topic_filter.split(',') if t.strip()]
        if topics:
            pool = pool.filter(Question.topic.in_(topics))

    pool = pool.all()

    if len(pool) == 0:
        return []

    n = min(config.total_questions, len(pool))
    selected = random.sample(pool, n)
    random.shuffle(selected)
    return selected


def calculate_exam_score(questions, user_answers_dict, config):
    """Calculate exam score with marking scheme."""
    correct = 0
    wrong = 0
    unattempted = 0
    bonus = 0
    marks = 0.0

    for question in questions:
        given = user_answers_dict.get(str(question.id), 'E').upper()
        if question.correct_ans == 'X':
            bonus += 1
            marks += config.marks_correct
        elif given in ['E', '', 'SKIP']:
            unattempted += 1
        elif given == question.correct_ans:
            correct += 1
            marks += config.marks_correct
        else:
            wrong += 1
            marks -= config.marks_wrong

    max_score = len(questions) * config.marks_correct
    percentage = round(marks / max_score * 100, 1) if max_score > 0 else 0
    if percentage < 0:
        percentage = 0

    return {
        'marks': round(marks, 2),
        'max_score': round(max_score, 2),
        'percentage': percentage,
        'correct': correct,
        'wrong': wrong,
        'unattempted': unattempted,
        'bonus': bonus
    }


def generate_sample_csv():
    """Generate sample CSV content for download."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['qno', 'section', 'topic', 'question', 'option_a', 'option_b',
                      'option_c', 'option_d', 'correct_ans', 'explanation'])
    writer.writerow([1, 1, 'Physics', 'What is the SI unit of force?',
                      'Newton', 'Joule', 'Watt', 'Pascal', 'A',
                      'Force is measured in Newton (N) in the SI system.'])
    writer.writerow([2, 1, 'Chemistry', 'What is the chemical formula of water?',
                      'H2O', 'CO2', 'NaCl', 'O2', 'A',
                      'Water is composed of two hydrogen atoms and one oxygen atom.'])
    writer.writerow([3, 2, 'Mathematics', 'What is the value of pi (approx)?',
                      '2.14', '3.14', '4.14', '1.14', 'B',
                      'Pi is approximately 3.14159...'])
    return output.getvalue()
